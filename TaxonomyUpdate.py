import openpyxl
import os
import requests
from bs4 import BeautifulSoup

# 1.0 Read urls from the excel file
# 1.1 Change the current path of the file
path = "C:\\Users\\maitreypatel\\Desktop\\TaxonomyUpdates\\LATTE-931-Pets-Taxonomy"
workBook = "Pets_Taxonomy_Update_V1.3.xlsx"
sheetNameRename = "Rename Validation"
sheetNameInsert = "Insert Validation"
sheetNameDelete = "Delete Validation"
sheetNameRedirect = "Redirect Validation"

def changeDirectory():
    changedir = os.chdir(path)
excelNames = []
getCrumb = []
# 1.2 Go to the excel file where the taxonomies to be validated
def getRenames():
    getRenames = openpyxl.load_workbook(workBook).get_sheet_by_name(sheetNameRename).columns[8]
    for renames in getRenames:
        names = renames.value
        req = requests.get(names)
        soup = BeautifulSoup(req.content, 'html.parser')
        crumbList=soup.find("div", {"id": "brd-crumbs"}).find_all("span")[0].text
        getCrumb.append(crumbList)
    #print(getCrumb)


# To test the rollback, change the column number
def getFunnyNames():
    getExcelNames = openpyxl.load_workbook(workBook).get_sheet_by_name(sheetNameRename).columns[10]
    for taxRenames in getExcelNames:
        excelNames.append(taxRenames.value)
    #print(excelNames)
    #return excelNames

#verify the rollback renames of the taxonomies


# Compare the parsed taxonomy values with the excel taxonomy values

def compareNames():
    notMatchedRenames = 0
    for lastCrumbs in getCrumb:
        if lastCrumbs not in excelNames:
            print("Crumbs - "+str(lastCrumbs) + " excel Names " + str(excelNames))
            notMatchedRenames= notMatchedRenames + 1
    print(notMatchedRenames)
    if notMatchedRenames==0:
        print("All renames are matching")

def verifyInserts():
    invalidInserts = 0
    getInsertNames = openpyxl.load_workbook(workBook).get_sheet_by_name(sheetNameInsert).columns[6]
    for inserts in getInsertNames:
        insertNames = inserts.value
        req_status = requests.get(insertNames).status_code
        if req_status != 200:
            print(insertNames)
            invalidInserts = invalidInserts + 1
    print(invalidInserts)
    if invalidInserts==0:
        print("All insert urls are correct")


def verifyDeletes():
    invalidDeleteCounts = 0
    getDeletedNames = openpyxl.load_workbook(workBook).get_sheet_by_name(sheetNameDelete).columns[6]


def redirects():
    numberofredirecturls = 0
    getRedirects = openpyxl.load_workbook(workBook).get_sheet_by_name(sheetNameDelete).columns[6]
    for redirecturls in getRedirects:
        redirectNames =  redirecturls.value
        redirectstatus = requests.get(redirecturls.rstrip('\n'), allow_redirects=True)
        if redirectstatus is not 301:
            print('Url is not redirecting - ' + str(redirecturls))
            numberofredirecturls = numberofredirecturls + 1
    if numberofredirecturls == 0:
       print('All urls are redirecting and there is ' + str(numberofredirecturls) + ' urls')
    else:
        print('number of Nonredirecting urls are ' + str(numberofredirecturls))



changeDirectory()
getRenames()
getFunnyNames()
compareNames()
verifyInserts()
redirects()







# Pets_Taxonomy_Update_V1.3.xls

# Save the taxonomies name from the excel file

# Parse the names of the taxonomies from the page sources


